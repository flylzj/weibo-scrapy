# -*- coding: utf-8 -*-
import openpyxl
import os
from weibo_scrapy.db import WeiboUser, Weibo, Base
from sqlalchemy.orm import sessionmaker
import datetime
from sqlalchemy import create_engine

# class WeiboPipeline(object):
#     def __init__(self):
#         if not os.path.exists('data'):
#             os.mkdir('data')
#         self.data_path = './data'
#         self.workbook_list = {}
#         self.title = [
#             '微博mid',
#             '微博用户id',
#             '用户昵称',
#             '微博内容',
#             '发布时间',
#             '转发数量',
#             '评论数量',
#             '点赞数量',
#             '转发源',
#             '微博用户性别',
#             '微博用户所在地',
#             '微博用户生日',
#             '微博用户简介',
#             '微博用户注册时间'
#         ]
#
#     def process_item(self, item, spider):
#         start_time = item.get('start_time')
#         if start_time not in self.workbook_list:
#             wb = openpyxl.Workbook()
#             self.workbook_list[start_time] = wb
#             sheet = wb.active
#             sheet.append(self.title)
#         else:
#             wb = self.workbook_list.get(start_time)
#             sheet = wb.active
#         # 说明这时间段爬取完成
#         if not item.get('end_time'):
#             spider.logger.info('finish' + item.get('start_time') + ".xlsx")
#             wb.save(os.path.join(self.data_path, item.get('start_time') + ".xlsx"))
#         else:
#             sheet.append([
#                 item.get('weibo_mid'),
#                 item.get('weibo_user_id'),
#                 item.get('weibo_username'),
#                 item.get('weibo_content'),
#                 item.get('weibo_pub_time'),
#                 item.get('weibo_post_count'),
#                 item.get('weibo_comment_count'),
#                 item.get('weibo_like_count'),
#                 item.get('weibo_from_url', ''),
#                 item.get('user_info').get('weibo_sexual', ''),
#                 item.get('user_info').get('weibo_location', ''),
#                 item.get('user_info').get('weibo_birth', ''),
#                 item.get('user_info').get('weibo_desc', ''),
#                 item.get('user_info').get('weibo_sign_up_date', '')
#             ])
#         return item
#
#     # def close_spider(self, spider):
#     #     for t, wb in self.workbook_list.items():
#     #         wb.save(os.path.join(self.data_path, t + ".xlsx"))


# 把数据存入excel的类
class WeiboExcelPipeline(object):
    def __init__(self):
        if not os.path.exists('data'):
            os.mkdir('data')
        self.data_path = './data'
        self.wb = openpyxl.Workbook()
        self.count = 0
        self.max = 10000
        self.data_index = 0
        self.title = [
            '微博mid',
            '微博用户id',
            '用户昵称',
            '微博内容',
            '发布时间',
            '转发数量',
            '评论数量',
            '点赞数量',
            '转发源',
            '微博用户性别',
            '微博用户所在地',
            '微博用户生日',
            '微博用户简介',
            '微博用户注册时间'
        ]

    def process_item(self, item, spider):
        sheet = self.wb.active
        if self.count == 0:
            sheet.append(self.title)
        sheet.append([
            item.get('weibo_mid'),
            item.get('weibo_user_id'),
            item.get('weibo_username'),
            item.get('weibo_content'),
            item.get('weibo_pub_time'),
            item.get('weibo_post_count'),
            item.get('weibo_comment_count'),
            item.get('weibo_like_count'),
            item.get('weibo_from_url', ''),
            item.get('user_info').get('weibo_sexual', ''),
            item.get('user_info').get('weibo_location', ''),
            item.get('user_info').get('weibo_birth', ''),
            item.get('user_info').get('weibo_desc', ''),
            item.get('user_info').get('weibo_sign_up_date', '')
        ])
        self.count += 1
        # 每10000条做一个新的excel
        if self.count > self.max:
            self.count = 0
            self.wb.save('data' + str(self.data_index))
            self.data_index += 1
            self.wb = openpyxl.Workbook()
        return item

    def close_spider(self, spider):
        self.wb.save(os.path.join(self.data_path, 'data.xlsx'))


# 把输入存入数据的类
class WeiboDbPipeline(object):
    # 从设置中导入mysql设置
    @classmethod
    def from_crawler(cls, crawler):
        settings = crawler.settings
        MYSQL_SETTING = settings.get('MYSQL_SETTING')
        return cls(MYSQL_SETTING)

    def __init__(self, MYSQL_SETTING):
        # 配置mysql
        # ENGINE = create_engine('sqlite:///./data.db')
        # 数据库引擎
        ENGINE =create_engine("mysql+pymysql://{}:{}@{}/{}".format(
            MYSQL_SETTING.get('user'),
            MYSQL_SETTING.get('password'),
            MYSQL_SETTING.get('host'),
            MYSQL_SETTING.get('db')
        ))
        self.ENGINE = ENGINE
        # 会话
        self.Session = sessionmaker(bind=self.ENGINE)
        # 初始化数据库
        # 自动建表
        self.init_table()

    def init_table(self):
        # 删除所有表，这里除非想要重新创建表，否者这一行保持注释
        # Base.metadata.drop_all(self.ENGINE)
        # 创建所有表
        Base.metadata.create_all(self.ENGINE)

    # 转换微博的日期字符串为python的date类型
    @staticmethod
    def convert_date(s, t):
        '''
        :param s: 2018年12月31日 22:54 / 1986年12月2日 / 9月9日 / 金牛座 / 2012-04-21
        :param t: birth / pub_date / sign_up_date
        :return:
        '''
        if t == 'birth':
            try:
                date = datetime.datetime.strptime(s, '%Y年%m月%d日')
            except Exception as e:
                date = datetime.date(year=1970, month=1, day=1)
        elif t == 'pub_date':
            if '秒前' in s:
                d = int(s.strip('秒前').strip())
                date = datetime.datetime.now() - datetime.timedelta(seconds=d + 10)
            elif '分钟前' in s:
                d = int(s.strip('分钟').strip())
                date = datetime.datetime.now() - datetime.timedelta(minutes=d)
            elif '今天' in s:
                s = s.strip('今天').strip()
                d = datetime.datetime.now()
                s = "{}年{}月{}日 ".format(d.year, d.month, d.day) + s
                date = datetime.datetime.strptime(s, '%Y年%m月%d日 %H:%M')
            elif '昨天' in s:
                s = s.strip('昨天').strip()
                d = datetime.datetime.now() - datetime.timedelta(days=1)
                s = "{}年{}月{}日 ".format(d.year, d.month, d.day) + s
                date = datetime.datetime.strptime(s, '%Y年%m月%d日 %H:%M')
            elif '年' not in s:
                d = datetime.datetime.now()
                s = "{}年".format(d.year) + s
                date = datetime.datetime.strptime(s, '%Y年%m月%d日 %H:%M')
            else:
                date = datetime.datetime.strptime(s, '%Y年%m月%d日 %H:%M')
        elif t == 'sign_up_date':
            date = datetime.datetime.strptime(s, '%Y-%m-%d')
        else:
            date = datetime.date(year=1970, month=1, day=1)
        return date

    # 通过mid查找数据库中的微博
    def search_weibo(self, mid):
        try:
            session = self.Session()
            res = session.query(Weibo).filter_by(
                mid=mid
            ).first()
            return res
        except Exception as e:
            return None

    # 通过user_id查找数据库中的user
    def search_weibo_user(self, user_id):
        try:
            session = self.Session()
            res = session.query(WeiboUser).filter_by(
                user_id=user_id
            ).first()
            return res
        except Exception as e:
            return None

    # 当item经过这个pipeline执行这个方法
    def process_item(self, item, spider):
        try:
            # 创建数据库session
            session = self.Session()
            # 如果数据库里没有这条记录就创建
            if not self.search_weibo_user(item.get('user_id')):
                weibo_user = WeiboUser(
                    user_id=item.get('weibo_user_id'),
                    username=item.get('weibo_username'),
                    sexual=item.get('user_info').get('weibo_sexual', ''),
                    location=item.get('user_info').get('weibo_location', ''),
                    raw_birth=item.get('user_info').get('weibo_birth', ''),
                    birth=self.convert_date(item.get('user_info').get('weibo_birth', ''), 'birth'),
                    desc=item.get('user_info').get('weibo_desc', ''),
                    raw_sign_up_date=item.get('user_info').get('weibo_sign_up_date', ''),
                    sign_up_date=self.convert_date(item.get('user_info').get('weibo_sign_up_date', ''), 'sign_up_date')
                )
                session.add(weibo_user)
                spider.logger.debug('add weibo_user {} success'.format(item.get('weibo_user_id')))
            # 如果数据库里没有这条记录就创建
            if not self.search_weibo_user(item.get('weibo_mid')):
                weibo = Weibo(
                    mid=item.get('weibo_mid'),
                    user_id=item.get('weibo_user_id'),
                    content=item.get('weibo_content'),
                    raw_pub_time=item.get('weibo_pub_time'),
                    pub_time=self.convert_date(item.get('weibo_pub_time'), 'pub_date'),
                    post_count=item.get('weibo_post_count'),
                    comment_count=item.get('weibo_comment_count'),
                    like_count=item.get('weibo_like_count'),
                    from_url=item.get('weibo_from_url', '')
                )
                session.add(weibo)
                spider.logger.debug('add weibo {} success'.format(item.get('weibo_mid')))
            session.commit()
            session.close()
        except Exception as e:
            spider.logger.error(e, exc_info=True)
        return item


